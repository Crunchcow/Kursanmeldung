from django.contrib import admin
from django.contrib.auth.admin import UserAdmin, GroupAdmin
from django.contrib.auth.models import User, Group
from django.utils.translation import gettext_lazy as _
from django.utils.html import format_html
from django.urls import reverse, path
from django.utils.http import urlencode
from django.http import HttpResponseRedirect, HttpResponse
from .models import Location, Course, CourseSession, Registration


# ---------------------------------------------------------------------------
# Benutzer- und Gruppen-Verwaltung ist gesperrt — läuft über ClubAuth
# ---------------------------------------------------------------------------

class ReadOnlyUserAdmin(UserAdmin):
    """Zeigt User an, erlaubt aber weder Anlage noch Bearbeitung noch Löschung."""
    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


class ReadOnlyGroupAdmin(GroupAdmin):
    """Zeigt Gruppen an, erlaubt aber weder Anlage noch Bearbeitung noch Löschung."""
    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


admin.site.unregister(User)
admin.site.unregister(Group)
admin.site.register(User, ReadOnlyUserAdmin)
admin.site.register(Group, ReadOnlyGroupAdmin)


@admin.register(Location)
class LocationAdmin(admin.ModelAdmin):
    list_display = ('name',)


class CourseSessionInline(admin.TabularInline):
    """Einzelne Kurseinheiten direkt am Kurs bearbeiten."""
    model = CourseSession
    extra = 0
    fields = ('date', 'is_cancelled', 'note')
    ordering = ('date',)


class RegistrationInline(admin.TabularInline):
    model = Registration
    extra = 0
    fields = ('first_name', 'last_name', 'email', 'phone', 'status', 'custom_price')
    readonly_fields = ('first_name', 'last_name', 'email', 'phone', 'status')
    show_change_link = True


@admin.register(CourseSession)
class CourseSessionAdmin(admin.ModelAdmin):
    """Direkte Verwaltung einzelner Einheiten (optional)."""
    list_display = ('course', 'date', 'is_cancelled', 'note')
    list_filter = ('course', 'is_cancelled')
    ordering = ('course', 'date')


@admin.register(Course)
class CourseAdmin(admin.ModelAdmin):
    list_display = (
        'name', 'course_type', 'start_date', 'end_date',
        'start_time', 'end_time', 'days',
        'utilization_display', 'registrations_link', 'attendance_export_link', 'is_closed',
    )
    list_editable = ('is_closed',)
    list_filter = ('is_closed', 'course_type', 'session_mode', 'days', 'locations', 'start_date', 'instructor_user')
    inlines = [CourseSessionInline, RegistrationInline]
    readonly_fields = ('session_count_display',)
    actions = ['export_attendance_list', 'generate_sessions_action', 'copy_course_with_participants']

    fieldsets = (
        (_('Allgemein'), {
            'fields': ('name', 'description', 'course_type', 'locations', 'instructor', 'instructor_user'),
        }),
        (_('Zeiten'), {
            'fields': ('start_date', 'end_date', 'start_time', 'end_time', 'days'),
        }),
        (_('Einheiten'), {
            'fields': ('session_mode', 'num_sessions', 'session_count_display'),
            'description': _(
                'Wähle den Modus und klicke dann die Aktion "Einheiten generieren". '
                'Im Modus "Manuell" können die Einheiten unten direkt eingetragen werden.'
            ),
        }),
        (_('Anmeldung'), {
            'fields': ('max_participants', 'price_member', 'price_non_member', 'allow_half', 'is_closed', 'publish_from'),
        }),
    )

    def utilization_display(self, obj):
        confirmed = obj.current_registrations()
        total = obj.max_participants
        pct = int(confirmed / total * 100) if total else 0
        if pct >= 100:
            color = '#c00000'
        elif pct >= 75:
            color = '#e67e00'
        else:
            color = '#2e7d32'
        return format_html(
            '<span style="color:{};font-weight:bold;">{}/{}</span> '
            '<span style="color:#888;font-size:0.85em;">({}&nbsp;%)</span>',
            color, confirmed, total, pct,
        )
    utilization_display.short_description = _('Auslastung')

    def registrations_link(self, obj):
        confirmed = obj.registration_set.filter(status='CONFIRMED').count()
        waitlist  = obj.registration_set.filter(status='WAITLIST').count()
        url = (
            reverse('admin:courses_registration_changelist')
            + '?' + urlencode({'course__id__exact': obj.pk})
        )
        label = f'{confirmed} Teilnehmer'
        if waitlist:
            label += f' + {waitlist} Warteliste'
        return format_html('<a href="{}">{} &rarr;</a>', url, label)
    registrations_link.short_description = _('Anmeldungen')
    registrations_link.admin_order_field = None

    def attendance_export_link(self, obj):
        url = reverse('admin:courses_course_export_attendance', args=[obj.pk])
        return format_html(
            '<a class="button" href="{}" '
            'style="background:#c00000;color:#fff;padding:3px 10px;'
            'border-radius:4px;white-space:nowrap;font-size:0.85em;'
            'text-decoration:none;">'
            '&#8595; Excel</a>',
            url
        )
    attendance_export_link.short_description = _('Anwesenheitsliste')
    attendance_export_link.admin_order_field = None

    change_list_template = 'admin/courses/course/change_list.html'

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                '<int:course_id>/export-attendance/',
                self.admin_site.admin_view(self.export_attendance_direct),
                name='courses_course_export_attendance',
            ),
            path(
                'archiv/',
                self.admin_site.admin_view(self.archive_view),
                name='courses_course_archiv',
            ),
        ]
        return custom + urls

    def archive_view(self, request):
        """Admin-Ansicht: abgelaufene Kurse im Archiv."""
        from datetime import date
        from django.shortcuts import render as django_render
        today = date.today()
        courses = (
            Course.objects
            .filter(end_date__lt=today)
            .order_by('-start_date')
        )
        context = {
            **self.admin_site.each_context(request),
            'title': 'Kursarchiv',
            'courses': courses,
            'opts': self.model._meta,
        }
        return django_render(request, 'admin/courses/course/archive.html', context)

    def export_attendance_direct(self, request, course_id):
        """Direkt-Download der Anwesenheitsliste fuer einen einzelnen Kurs."""
        from django.shortcuts import get_object_or_404
        course = get_object_or_404(Course, pk=course_id)
        if (
            request.user.groups.filter(name='Kursleitung').exists()
            and course.instructor_user != request.user
        ):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied
        qs = Course.objects.filter(pk=course_id)
        return self.export_attendance_list(request, qs)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.groups.filter(name='Kursleitung').exists():
            return qs.filter(instructor_user=request.user)
        return qs

    def has_add_permission(self, request):
        if request.user.groups.filter(name='Kursleitung').exists():
            return False
        return super().has_add_permission(request)

    def has_change_permission(self, request, obj=None):
        if request.user.groups.filter(name='Kursleitung').exists():
            if obj is None:
                return True
            return obj.instructor_user == request.user
        return super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if request.user.groups.filter(name='Kursleitung').exists():
            return False
        return super().has_delete_permission(request, obj)

    def session_count_display(self, obj):
        count = obj.session_count()
        db_count = obj.sessions.count()
        if db_count:
            return format_html(
                '{} Einheiten ({} generiert, {} aktiv)',
                count, db_count, count,
            )
        return f'{count} Einheiten (noch nicht generiert)'
    session_count_display.short_description = _('Einheiten')

    def generate_sessions_action(self, request, queryset):
        """Generiert Kurseinheiten fuer alle ausgewaehlten Kurse basierend auf dem Modus."""
        from django.contrib import messages as msg
        total = 0
        for course in queryset:
            course.generate_sessions(skip_holidays=True)
            total += course.session_count()
        self.message_user(
            request,
            _(f'{queryset.count()} Kurs(e) neu generiert, insgesamt {total} Einheiten.'),
        )
    generate_sessions_action.short_description = _('Einheiten generieren (NRW-Feiertage überspringen)')

    def copy_course_with_participants(self, request, queryset):
        """Kopiert einen Kurs inkl. aller aktiven Teilnehmer (als Warteliste)."""
        import uuid as uuid_lib
        if queryset.count() != 1:
            self.message_user(
                request,
                _('Bitte genau einen Kurs auswählen.'),
                level='error',
            )
            return

        original = queryset.first()

        # Kurs duplizieren
        new_course = Course(
            name=f'{original.name} (Folgekurs)',
            description=original.description,
            start_date=None,   # muss vom Admin gesetzt werden
            end_date=None,
            start_time=original.start_time,
            end_time=original.end_time,
            days=original.days,
            max_participants=original.max_participants,
            price_member=original.price_member,
            price_non_member=original.price_non_member,
            allow_half=original.allow_half,
            is_closed=True,  # geschlossen bis Admin alles eingerichtet hat
            instructor=original.instructor,
            instructor_user=original.instructor_user,
            session_mode=original.session_mode,
            num_sessions=original.num_sessions,
            course_type=original.course_type,
        )
        new_course.save()
        new_course.locations.set(original.locations.all())

        # Alle aktiven Teilnehmer uebernehmen (CANCELLED ausschliessen)
        source_regs = original.registration_set.exclude(status='CANCELLED')
        new_regs = []
        for reg in source_regs:
            new_regs.append(Registration(
                course=new_course,
                first_name=reg.first_name,
                last_name=reg.last_name,
                email=reg.email,
                phone=reg.phone,
                iban=reg.iban,
                bic=reg.bic,
                account_holder=reg.account_holder,
                terms_accepted=reg.terms_accepted,
                is_member=reg.is_member,
                half_course=reg.half_course,
                status='WAITLIST',  # alle erstmal auf Warteliste
                cancel_token=uuid_lib.uuid4(),
            ))
        Registration.objects.bulk_create(new_regs)

        self.message_user(
            request,
            _(f'Folgekurs "{new_course.name}" angelegt mit {len(new_regs)} Teilnehmern auf der '
              f'Warteliste. Bitte Datum setzen, unerwünschte Teilnehmer löschen und '
              f'dann "Bestätigen und Info-Mail senden" ausführen.'),
        )
        return HttpResponseRedirect(
            reverse('admin:courses_course_change', args=[new_course.pk])
        )
    copy_course_with_participants.short_description = _('Kurs kopieren (Folgekurs mit allen Teilnehmern als Warteliste)')

    def export_attendance_list(self, request, queryset):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        import zipfile, io

        red_fill   = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
        grey_fill  = PatternFill(start_color='D8D8D8', end_color='D8D8D8', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        thin       = Side(style='thin')
        border     = Border(left=thin, right=thin, top=thin, bottom=thin)
        center     = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        def make_wb(course):
            wb = Workbook()
            ws = wb.active
            ws.title = 'Anwesenheit'

            dates     = course.session_dates()
            locations = ', '.join(l.name for l in course.locations.all()) or '-'
            days_str  = ', '.join(course.days) if course.days else '-'
            last_col  = get_column_letter(3 + len(dates))

            ws.merge_cells(f'A1:{last_col}1')
            c = ws['A1']
            c.value     = f'Anwesenheitsliste \u2013 {course.name}'
            c.font      = Font(color='FFFFFF', bold=True, size=14)
            c.fill      = red_fill
            c.alignment = center
            ws.row_dimensions[1].height = 24

            ws.merge_cells(f'A2:{last_col}2')
            c = ws['A2']
            start_str = course.start_date.strftime('%d.%m.%Y') if course.start_date else '-'
            end_str   = course.end_date.strftime('%d.%m.%Y')   if course.end_date   else '-'
            c.value = (
                f'Zeitraum: {start_str} \u2013 {end_str}   |   '
                f'Zeit: {course.start_time.strftime("%H:%M")} \u2013 '
                f'{course.end_time.strftime("%H:%M")} Uhr   |   '
                f'Tage: {days_str}   |   Ort: {locations}   |   '
                f'Einheiten: {len(dates)}'
            )
            c.font      = Font(size=10)
            c.fill      = grey_fill
            c.alignment = left_align
            ws.row_dimensions[2].height = 16

            ws.row_dimensions[3].height = 8

            for col, h in enumerate(['Nr.', 'Nachname', 'Vorname'], 1):
                c = ws.cell(row=4, column=col, value=h)
                c.font      = Font(color='FFFFFF', bold=True, size=11)
                c.fill      = red_fill
                c.border    = border
                c.alignment = center

            for i, d in enumerate(dates, 1):
                c = ws.cell(row=4, column=3 + i, value=d.strftime('%d.%m.\n%a'))
                c.font      = Font(color='FFFFFF', bold=True, size=10)
                c.fill      = red_fill
                c.border    = border
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[4].height = 30

            registrations = (
                course.registration_set
                .filter(status='CONFIRMED')
                .order_by('last_name', 'first_name')
            )
            for idx, reg in enumerate(registrations, 1):
                data_row  = 4 + idx
                row_fill  = white_fill if idx % 2 == 1 else grey_fill

                for col, val in enumerate([idx, reg.last_name, reg.first_name], 1):
                    c = ws.cell(row=data_row, column=col, value=val)
                    c.border    = border
                    c.fill      = row_fill
                    c.alignment = center if col == 1 else left_align

                for i in range(1, len(dates) + 1):
                    c = ws.cell(row=data_row, column=3 + i, value='')
                    c.border    = border
                    c.fill      = row_fill
                    c.alignment = center

                ws.row_dimensions[data_row].height = 18

            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 22
            ws.column_dimensions['C'].width = 16
            for i in range(1, len(dates) + 1):
                ws.column_dimensions[get_column_letter(3 + i)].width = 7

            ws.freeze_panes = 'D5'
            return wb

        courses = list(queryset)
        if len(courses) == 1:
            wb = make_wb(courses[0])
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            safe = courses[0].name.replace('/', '-')
            response['Content-Disposition'] = f'attachment; filename="Anwesenheit_{safe}.xlsx"'
            wb.save(response)
            return response

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            for course in courses:
                xlsx = io.BytesIO()
                make_wb(course).save(xlsx)
                safe = course.name.replace('/', '-')
                zf.writestr(f'Anwesenheit_{safe}.xlsx', xlsx.getvalue())
        buf.seek(0)
        response = HttpResponse(buf, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="Anwesenheitslisten.zip"'
        return response

    export_attendance_list.short_description = str(_("Anwesenheitsliste als Excel exportieren"))


@admin.register(Registration)
class RegistrationAdmin(admin.ModelAdmin):
    list_display = ('course', 'last_name', 'first_name', 'email', 'phone', 'status', 'custom_price_display')
    list_filter = ('status', 'course')
    search_fields = ('first_name', 'last_name', 'email', 'phone')
    readonly_fields = ('created', 'cancel_token', 'waitlist_position_display')
    fields = (
        'course', 'status',
        'first_name', 'last_name', 'email', 'phone',
        'iban', 'bic', 'account_holder',
        'is_member', 'half_course', 'custom_price',
        'terms_accepted', 'created', 'cancel_token', 'waitlist_position_display',
    )
    actions = ['export_as_csv', 'export_debits', 'export_wiso_meinverein', 'confirm_and_notify']

    def custom_price_display(self, obj):
        if obj.custom_price is not None:
            return format_html(
                '<span style="color:#e67e00;font-weight:bold;">{} &euro; (individuell)</span>',
                obj.custom_price,
            )
        return format_html('<span style="color:#888;">{} &euro;</span>', obj.price())
    custom_price_display.short_description = _('Zu zahlender Betrag')

    def waitlist_position_display(self, obj):
        pos = obj.waitlist_position()
        if pos is None:
            return '-'
        return f'Platz {pos} auf der Warteliste'
    waitlist_position_display.short_description = _('Wartelisten-Position')

    def changelist_view(self, request, extra_context=None):
        if (
            request.user.groups.filter(name='Kursleitung').exists()
            and 'course__id__exact' not in request.GET
        ):
            return HttpResponseRedirect(reverse('admin:courses_course_changelist'))
        return super().changelist_view(request, extra_context)

    def get_actions(self, request):
        actions = super().get_actions(request)
        is_kassierer = request.user.groups.filter(name='Kassierer').exists()
        if not (request.user.is_superuser or is_kassierer):
            actions.pop('export_wiso_meinverein', None)
        return actions

    def confirm_and_notify(self, request, queryset):
        """Setzt ausgewaehlte WAITLIST-Anmeldungen auf CONFIRMED und sendet Info-Mail."""
        from .models import _send_waitlist_promotion_email
        count = 0
        for reg in queryset.filter(status='WAITLIST'):
            reg.status = 'CONFIRMED'
            reg.save(update_fields=['status'])
            _send_waitlist_promotion_email(reg)
            count += 1
        self.message_user(
            request,
            _(f'{count} Anmeldung(en) bestätigt und Info-Mail gesendet.'),
        )
    confirm_and_notify.short_description = _('Auswahl bestätigen + Info-Mail senden (Folgekurs)')

    def export_as_csv(self, request, queryset):
        import csv
        field_names = ['course', 'first_name', 'last_name', 'email', 'phone', 'status', 'terms_accepted', 'price', 'created']
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=registrations.csv'
        writer = csv.writer(response)
        writer.writerow(field_names)
        for obj in queryset:
            writer.writerow([
                obj.course.name, obj.first_name, obj.last_name, obj.email, obj.phone,
                obj.status, obj.terms_accepted, obj.price(), obj.created,
            ])
        return response
    export_as_csv.short_description = str(_('Anmeldungen als CSV exportieren'))

    def export_debits(self, request, queryset):
        """CSV fuer den Kassierer, nur CONFIRMED."""
        import csv
        headers = [
            str(_('Kurs')), str(_('Vorname')), str(_('Nachname')),
            str(_('IBAN')), str(_('BIC')), str(_('Kontoinhaber')), str(_('Betrag')),
        ]
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=einzuege.csv'
        writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        for reg in queryset.filter(status='CONFIRMED'):
            amount = '{:.2f}'.format(reg.total_price()).replace('.', ',')
            writer.writerow([
                reg.course.name, reg.first_name, reg.last_name,
                reg.iban, reg.bic or '', reg.account_holder, amount,
            ])
        return response
    export_debits.short_description = _('Einzüge als CSV exportieren')

    def export_wiso_meinverein(self, request, queryset):
        """Exportiert NUR bestaetigte Anmeldungen als WISO-MeinVerein-CSV."""
        from django.contrib import messages as msg
        import csv

        is_kassierer = request.user.groups.filter(name='Kassierer').exists()
        if not (request.user.is_superuser or is_kassierer):
            self.message_user(request, _('Sie haben keine Berechtigung für diesen Export.'), msg.ERROR)
            return

        headers = [
            'Vorname', 'Nachname', 'IBAN', 'BIC', 'Kontoinhaber',
            'Betrag', 'Verwendungszweck', 'Mandatsreferenz', 'Mandatsdatum',
        ]
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename=wiso_meinverein_lastschriften.csv'
        writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)

        # Nur CONFIRMED exportieren
        for reg in queryset.filter(status='CONFIRMED'):
            amount       = '{:.2f}'.format(reg.total_price()).replace('.', ',')
            mandate_date = reg.created.strftime('%d.%m.%Y')
            mandate_ref  = f'KURS-{reg.id:06d}'
            purpose_parts = [reg.course.name]
            if reg.half_course and reg.course.allow_half:
                purpose_parts.append('(Halber Kurs)')
            purpose = ' '.join(purpose_parts)

            writer.writerow([
                reg.first_name, reg.last_name, reg.iban, reg.bic or '',
                reg.account_holder, amount, purpose, mandate_ref, mandate_date,
            ])
        return response
    export_wiso_meinverein.short_description = _('WISO MeinVerein – SEPA-Lastschriften exportieren (nur Bestätigt)')

    def has_module_permission(self, request):
        if request.user.groups.filter(name__in=['Kursleitung', 'Kassierer']).exists():
            return True
        return super().has_module_permission(request)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if request.user.groups.filter(name='Kursleitung').exists():
            return qs.filter(course__instructor_user=request.user)
        return qs

    def has_view_permission(self, request, obj=None):
        if request.user.groups.filter(name__in=['Kursleitung', 'Kassierer']).exists():
            return True
        return super().has_view_permission(request, obj)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        if request.user.groups.filter(name__in=['Kursleitung', 'Kassierer']).exists():
            return False if obj else True
        return super().has_change_permission(request, obj)

    def has_delete_permission(self, request, obj=None):
        if request.user.groups.filter(name__in=['Kursleitung', 'Kassierer']).exists():
            return False
        return super().has_delete_permission(request, obj)
